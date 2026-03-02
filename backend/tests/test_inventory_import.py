"""
End-to-end tests for the inventory bulk import feature.
Covers template download, upload of the downloaded template (exact format),
custom template-format xlsx, plain csv, plain xlsx, and error cases.
"""
import io
import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlmodel import Session, select

from app.database import engine
from app.models import Product
from main import app as fastapi_app


# ── helpers ───────────────────────────────────────────────────────────────────

def _make_template_xlsx(rows: list[list]) -> bytes:
    """Build an xlsx in the exact DukaPOS template format:
       row 1 = title banner, row 2 = column headers, row 3 = hints, rows 4+ = data.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["DukaPOS -- Inventory Import Template (* = required)"])
    ws.append([
        "Item Name *", "Barcode *", "Selling Price *",
        "Buying Price", "Current Stock", "Low Stock Limit",
        "Wholesale Price", "Wholesale Threshold", "Category",
    ])
    ws.append(["hint-name", "hint-bc", "hint-sell", "hint-buy",
               "hint-stk", "hint-min", "hint-wp", "hint-wt", "hint-cat"])
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_plain_xlsx(headers: list, rows: list[list]) -> bytes:
    """Build a plain xlsx with headers in row 1 and data from row 2."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _product_exists(barcode: str) -> bool:
    with Session(engine) as s:
        return s.exec(select(Product).where(Product.barcode == barcode)).first() is not None


def _delete_barcodes(*barcodes: str):
    with Session(engine) as s:
        for bc in barcodes:
            p = s.exec(select(Product).where(Product.barcode == bc)).first()
            if p:
                s.delete(p)
        s.commit()


# ── fixtures ──────────────────────────────────────────────────────────────────

IMP_BCS = ["IMP-001", "IMP-002", "IMP-003", "IMP-004", "IMP-005"]


@pytest.fixture(autouse=True)
def cleanup_import_products():
    _delete_barcodes(*IMP_BCS)
    yield
    _delete_barcodes(*IMP_BCS)


# ── tests ─────────────────────────────────────────────────────────────────────

def test_template_download(client: TestClient):
    """GET /inventory/template returns a valid xlsx file."""
    r = client.get("/inventory/template")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers.get("content-type", "")
    # Must be a valid xlsx (openpyxl can open it)
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb.active
    # Row 2 must contain the column headers (9 columns now)
    headers = [str(ws.cell(2, c).value or "").strip() for c in range(1, 10)]
    assert "Item Name *" in headers
    assert "Barcode *" in headers
    assert "Selling Price *" in headers
    assert "Category" in headers


def test_upload_downloaded_template_unchanged(client: TestClient):
    """Upload the exact file returned by /inventory/template — must succeed (not 'Missing columns')."""
    template_bytes = client.get("/inventory/template").content
    r = client.post(
        "/inventory/upload",
        files={"file": ("inventory_import_template.xlsx", io.BytesIO(template_bytes),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200, f"Expected 200, got {r.status_code}: {r.text[:400]}"
    data = r.json()
    # The 10 sample rows should all have been created/updated (no column-level errors)
    assert data["created"] + data["updated"] == 10, data
    assert data.get("errors", []) == [], data.get("errors")


def test_upload_template_format_custom_rows(client: TestClient):
    """Template-format xlsx with our own product rows imports correctly."""
    xlsx = _make_template_xlsx([
        ["Widget Alpha", "IMP-001", 100.0, 70.0, 50, 5, None, None, "Electronics"],
        ["Widget Beta",  "IMP-002", 250.0, 180.0, 30, 3, 230.0, 12, "General"],
    ])
    r = client.post(
        "/inventory/upload",
        files={"file": ("custom.xlsx", io.BytesIO(xlsx),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200, r.text[:300]
    data = r.json()
    assert data["created"] + data["updated"] == 2, data
    assert data.get("errors", []) == [], data.get("errors")
    assert _product_exists("IMP-001")
    assert _product_exists("IMP-002")
    # Verify prices were stored correctly
    with Session(engine) as s:
        p = s.exec(select(Product).where(Product.barcode == "IMP-001")).first()
        assert p is not None
        assert p.price_selling == 100.0
        assert p.price_buying == 70.0
        assert p.stock_quantity == 50
        p2 = s.exec(select(Product).where(Product.barcode == "IMP-002")).first()
        assert p2 is not None
        assert p2.wholesale_price == 230.0
        assert p2.wholesale_threshold == 12


def test_upload_template_update_existing(client: TestClient):
    """Re-importing an existing barcode updates the product, does not duplicate."""
    xlsx1 = _make_template_xlsx([["Widget Gamma", "IMP-003", 100.0, 70.0, 20, 5, None, None, "Clothing"]])
    r1 = client.post("/inventory/upload",
        files={"file": ("c1.xlsx", io.BytesIO(xlsx1),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r1.json()["created"] == 1

    xlsx2 = _make_template_xlsx([["Widget Gamma UPDATED", "IMP-003", 150.0, 90.0, 99, 5, None, None, "Clothing"]])
    r2 = client.post("/inventory/upload",
        files={"file": ("c2.xlsx", io.BytesIO(xlsx2),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r2.status_code == 200, r2.text[:300]
    d2 = r2.json()
    assert d2["created"] == 0 and d2["updated"] == 1, d2
    assert d2.get("errors", []) == []
    with Session(engine) as s:
        p = s.exec(select(Product).where(Product.barcode == "IMP-003")).first()
        assert p is not None
        assert p.price_selling == 150.0
        assert p.stock_quantity == 99


def test_upload_plain_csv(client: TestClient):
    """Plain CSV with standard headers (no title/hint rows) imports correctly."""
    csv_content = (
        b"name,barcode,price_selling,price_buying,stock_quantity\n"
        b"CSV Widget,IMP-004,99.0,60.0,10\n"
    )
    r = client.post("/inventory/upload",
        files={"file": ("plain.csv", io.BytesIO(csv_content), "text/csv")})
    assert r.status_code == 200, r.text[:300]
    d = r.json()
    assert d["created"] + d["updated"] == 1, d
    assert d.get("errors", []) == []
    assert _product_exists("IMP-004")


def test_upload_plain_xlsx_no_title(client: TestClient):
    """Plain Excel file (headers in row 1, data from row 2) imports correctly."""
    xlsx = _make_plain_xlsx(
        ["Item Name", "Barcode", "Selling Price", "Buying Price", "Current Stock"],
        [["Plain Product", "IMP-005", 150.0, 100.0, 20]],
    )
    r = client.post("/inventory/upload",
        files={"file": ("plain.xlsx", io.BytesIO(xlsx),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200, r.text[:300]
    d = r.json()
    assert d["created"] + d["updated"] == 1, d
    assert d.get("errors", []) == []
    assert _product_exists("IMP-005")


def test_upload_wrong_extension_rejected(client: TestClient):
    """Non-.xlsx/.csv file returns 400."""
    r = client.post("/inventory/upload",
        files={"file": ("data.txt", io.BytesIO(b"hello"), "text/plain")})
    assert r.status_code == 400
    assert "xlsx" in r.json()["detail"].lower() or "csv" in r.json()["detail"].lower()


def test_upload_missing_required_columns(client: TestClient):
    """CSV without barcode column returns 400 with descriptive error."""
    csv_bad = b"name,price_selling\nBad Product,50.0\n"
    r = client.post("/inventory/upload",
        files={"file": ("bad.csv", io.BytesIO(csv_bad), "text/csv")})
    assert r.status_code == 400
    detail = r.json()["detail"].lower()
    assert "barcode" in detail
    assert "missing" in detail


def test_upload_optional_columns_have_defaults(client: TestClient):
    """CSV with only name+barcode+price_selling: optional fields get defaults."""
    csv_min = b"name,barcode,price_selling\nMinimal Product,IMP-001,80.0\n"
    r = client.post("/inventory/upload",
        files={"file": ("min.csv", io.BytesIO(csv_min), "text/csv")})
    assert r.status_code == 200, r.text[:300]
    d = r.json()
    assert d["created"] + d["updated"] == 1
    with Session(engine) as s:
        p = s.exec(select(Product).where(Product.barcode == "IMP-001")).first()
        assert p is not None
        assert p.stock_quantity == 0
        assert p.min_stock_alert == 5
        assert p.price_buying == 80.0  # defaults to price_selling
        assert p.wholesale_price is None
