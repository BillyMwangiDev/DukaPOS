"""Inventory bulk upload (Excel/CSV) and stock adjustment."""
import io
from datetime import datetime, timezone
from typing import List, Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlmodel import Session, select
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.database import engine
from app.models import Product, StockAdjustment
from app.websocket_manager import broadcast_sync, create_event, EventType

router = APIRouter(prefix="/inventory", tags=["inventory"])


@router.get("/template")
def download_template():
    """Return a pre-formatted Excel template for bulk inventory import."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"

    HEADER_BG = "1E3A5F"
    HINT_BG = "EBF1F8"
    HINT_FG = "5A6A7A"
    BORDER_COL = "C8D4E0"
    ODD_ROW = "F9FBFD"
    EVEN_ROW = "FFFFFF"

    thin = Side(style="thin", color=BORDER_COL)
    cb = Border(left=thin, right=thin, top=thin, bottom=thin)

    columns = [
        ("Item Name *", 28, "Full product name, e.g. Brookside Milk 500ml"),
        ("Barcode *", 18, "Unique barcode / SKU (digits or text)"),
        ("Selling Price *", 16, "VAT-inclusive retail price in KSh"),
        ("Buying Price", 16, "Your cost / buying price in KSh (optional)"),
        ("Current Stock", 15, "Units in stock right now (default: 0)"),
        ("Low Stock Limit", 16, "Alert when stock falls below this (default: 5)"),
        ("Wholesale Price", 16, "Bulk price if different from retail (optional)"),
        ("Wholesale Threshold", 20, "Min qty to qualify for wholesale price"),
    ]
    # Row 1: title banner
    ws.merge_cells(f"A1:{get_column_letter(len(columns))}1")
    tc = ws["A1"]
    tc.value = "DukaPOS — Inventory Import Template    (* = required)"
    tc.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    tc.fill = PatternFill("solid", fgColor="0D2137")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Row 2: column headers
    for ci, (label, width, _) in enumerate(columns, 1):
        required = label.endswith(" *")
        c = ws.cell(row=2, column=ci, value=label)
        c.font = Font(name="Calibri", bold=True, size=11, color="FFD700" if required else "FFFFFF")
        c.fill = PatternFill("solid", fgColor=HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = cb
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[2].height = 28

    # Row 3: hint row
    for ci, (_, _, hint) in enumerate(columns, 1):
        c = ws.cell(row=3, column=ci, value=hint)
        c.font = Font(name="Calibri", italic=True, size=9, color=HINT_FG)
        c.fill = PatternFill("solid", fgColor=HINT_BG)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = cb
    ws.row_dimensions[3].height = 20

    # Sample rows
    samples = [
        ("Brookside Milk 500ml", "6001059039614", 60.0, 48.0, 24, 5, None, None),
        ("Unga Dola 2kg", "6001000105001", 165.0, 130.0, 50, 10, 155.0, 10),
        ("Royco Mchuzi Mix 75g", "5900617024152", 45.0, 36.0, 100, 20, 40.0, 20),
        ("Indomie Noodles Chicken 70g", "8996001300406", 25.0, 18.0, 200, 30, 22.0, 50),
        ("Ketepa Pride Tea Bags 25s", "6001000501024", 85.0, 65.0, 60, 12, 80.0, 10),
        ("Jogoo Maize Flour 2kg", "6001000301001", 155.0, 120.0, 75, 10, 148.0, 10),
        ("Kimbo Cooking Fat 500g", "6001000601001", 175.0, 140.0, 35, 8, 168.0, 5),
        ("Nescafe Classic 50g", "7613036961752", 420.0, 330.0, 20, 4, None, None),
        ("Simba Chips Salt & Vinegar 36g", "6001106001243", 40.0, 30.0, 150, 24, 35.0, 36),
        ("Softcare Baby Wipes 80s", "6009174200103", 220.0, 175.0, 25, 5, 210.0, 5),
    ]
    price_fmt = "#,##0.00"

    for ri, (name, barcode, sell, buy, stock, min_s, wp, wt) in enumerate(samples, 4):
        rf = PatternFill("solid", fgColor=ODD_ROW if ri % 2 == 1 else EVEN_ROW)
        data = [name, barcode, sell, buy, stock, min_s, wp, wt]
        fmts = [None, None, price_fmt, price_fmt, "0", "0", price_fmt, "0"]
        for ci, (val, fmt) in enumerate(zip(data, fmts), 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = rf
            c.border = cb
            c.font = Font(name="Calibri", size=10)
            c.alignment = Alignment(vertical="center")
            if fmt and val is not None:
                c.number_format = fmt
        ws.row_dimensions[ri].height = 18

    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="inventory_import_template.xlsx"'},
    )


def get_db():
    with Session(engine) as session:
        yield session


# ── Stock Adjustment ──────────────────────────────────────────────────────────

class StockAdjustCreate(BaseModel):
    product_id: int
    quantity_change: int  # positive = add, negative = remove
    reason: str  # "Damage", "Expired", "Theft", "Received", "Correction"
    staff_id: Optional[int] = None


class StockAdjustRead(BaseModel):
    id: int
    product_id: int
    staff_id: Optional[int]
    quantity_change: int
    reason: str
    timestamp: str  # ISO format

    model_config = {"from_attributes": True}


@router.post("/adjust", response_model=StockAdjustRead, status_code=201)
def adjust_stock(data: StockAdjustCreate, session: Session = Depends(get_db)):
    """Create a stock adjustment (updates product stock_quantity and logs the change)."""
    if data.quantity_change == 0:
        raise HTTPException(status_code=400, detail="quantity_change cannot be zero")
    if abs(data.quantity_change) > 100_000:
        raise HTTPException(status_code=400, detail="quantity_change exceeds maximum allowed (100,000)")
    product = session.get(Product, data.product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    new_qty = product.stock_quantity + data.quantity_change
    if new_qty < 0:
        raise HTTPException(
            status_code=400,
            detail=f"Adjustment would result in negative stock ({new_qty}). Current: {product.stock_quantity}",
        )
    product.stock_quantity = new_qty
    session.add(product)
    adj = StockAdjustment(
        product_id=data.product_id,
        staff_id=data.staff_id,
        quantity_change=data.quantity_change,
        reason=data.reason,
        timestamp=datetime.now(timezone.utc).replace(tzinfo=None),
    )
    session.add(adj)
    session.commit()
    # Broadcast updated stock to all connected terminals
    broadcast_sync(create_event(
        EventType.INVENTORY_UPDATED,
        {"product_id": product.id, "product_name": product.name, "new_quantity": product.stock_quantity},
    ))
    session.refresh(adj)
    return StockAdjustRead(
        id=adj.id,
        product_id=adj.product_id,
        staff_id=adj.staff_id,
        quantity_change=adj.quantity_change,
        reason=adj.reason,
        timestamp=adj.timestamp.isoformat() + "Z",
    )


@router.get("/adjustments", response_model=List[StockAdjustRead])
def list_adjustments(
    product_id: Optional[int] = Query(None),
    limit: int = Query(50, le=200),
    session: Session = Depends(get_db),
):
    """List recent stock adjustments, optionally filtered by product."""
    stmt = select(StockAdjustment).order_by(StockAdjustment.timestamp.desc()).limit(limit)  # type: ignore[attr-defined]
    if product_id is not None:
        stmt = select(StockAdjustment).where(StockAdjustment.product_id == product_id).order_by(StockAdjustment.timestamp.desc()).limit(limit)  # type: ignore[attr-defined]
    rows = session.exec(stmt).all()
    return [
        StockAdjustRead(
            id=r.id,
            product_id=r.product_id,
            staff_id=r.staff_id,
            quantity_change=r.quantity_change,
            reason=r.reason,
            timestamp=r.timestamp.isoformat() + "Z",
        )
        for r in rows
    ]


# Expected column names (case-insensitive, strip); map to Product fields
COLUMN_ALIASES = {
    "item name": "name",
    "name": "name",
    "product name": "name",
    "product_name": "name",
    "barcode": "barcode",
    "code": "barcode",  # API/test compatibility
    "buying price": "price_buying",
    "buyingprice": "price_buying",
    "cost": "price_buying",
    "selling price": "price_selling",
    "sellingprice": "price_selling",
    "price": "price_selling",
    "current stock": "stock_quantity",
    "stock": "stock_quantity",
    "quantity": "stock_quantity",
    "low stock limit": "min_stock_alert",
    "min stock": "min_stock_alert",
    "min_stock_alert": "min_stock_alert",
    "wholesale price": "wholesale_price",
    "wholesaleprice": "wholesale_price",
    "wholesale_price": "wholesale_price",
    "wholesale threshold": "wholesale_threshold",
    "wholesalethreshold": "wholesale_threshold",
    "wholesale_threshold": "wholesale_threshold",
}


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize column names to Product field names where possible."""
    out = {}
    for c in df.columns:
        key = str(c).strip().lower()
        if key.endswith(" *"):   # strip required-field marker from template headers
            key = key[:-2].strip()
        if key in COLUMN_ALIASES:
            out[COLUMN_ALIASES[key]] = df[c]
        else:
            out[key] = df[c]
    return pd.DataFrame(out)


# All column-label variants that hold barcodes — passed to pd.read_excel converters
# so that numeric cells (e.g. 6001059039614.0 or 1.23E+13) become clean strings.
_BARCODE_LABELS = {"Barcode *", "Barcode", "barcode", "code", "Code"}


def _barcode_converter(val) -> str:
    """Convert a raw Excel barcode cell to a clean string.

    Handles:
      • float encoding  — 6001059039614.0 → "6001059039614"
      • scientific notation — 1.2345678901234e+13 → "12345678901234"
      • None / NaN  → ""
    Note: leading zeros lost by Excel (e.g. "00440…" stored as 440…) cannot
    be recovered here — users should format barcode columns as Text in Excel.
    """
    if val is None:
        return ""
    if isinstance(val, float):
        if val != val:  # NaN is the only value where x != x
            return ""
        return str(int(val))
    return str(val).strip()


# Keywords that identify a row as a column-header row (matched case-insensitively, exact).
# Two or more of these must be present in the same row for it to be treated as the header.
_HEADER_KEYWORDS: set[str] = {
    "item name", "item name *",
    "name", "product name", "product_name",
    "barcode", "barcode *", "code", "sku",
    "selling price", "selling price *", "price", "price_selling",
    "buying price", "buying price *", "cost",
}


def _find_header_row(df_all: pd.DataFrame) -> Optional[int]:
    """Scan the first 5 rows to find the one that looks like column headers.

    Returns the 0-based row index, or None if no recognisable header is found.
    """
    for i in range(min(5, len(df_all))):
        row_lower: set[str] = set()
        for v in df_all.iloc[i]:
            if v is not None and not (isinstance(v, float) and pd.isna(v)):
                row_lower.add(str(v).strip().lower())
        if len(row_lower & _HEADER_KEYWORDS) >= 2:
            return i
    return None


def _read_upload(file: UploadFile) -> tuple[pd.DataFrame, int]:
    """Read .xlsx or .csv into a normalised DataFrame.

    Returns ``(df, data_start_row)`` where *data_start_row* is the 1-based
    Excel row number of the first data row (used to build human-readable
    per-row error messages).

    Handles three layouts automatically:
      • Plain file  — headers in row 0, data from row 1 (data_start_row = 2).
      • DukaPOS template — title banner (row 0), headers (row 1),
        hints/description row (row 2), data from row 3 (data_start_row = 4).
      • Any file where headers are offset — detected by scanning the first
        five rows for known column-header keywords.

    Barcode columns are read with a converter that normalises float-encoded
    values (e.g. ``6001059039614.0`` or ``1.23E+13``) to clean strings.
    """
    raw = file.file.read()
    converters = {label: _barcode_converter for label in _BARCODE_LABELS}

    if file.filename and file.filename.lower().endswith(".csv"):
        df = pd.read_csv(
            io.BytesIO(raw),
            converters=converters,
            na_values=["", "N/A", "n/a", "-"],
            keep_default_na=True,
        )
        return df, 2  # row 1 = header, row 2 = first data row

    # ── XLSX: scan every row to locate the real header row ───────────────────
    df_all = pd.read_excel(io.BytesIO(raw), header=None)
    header_row_idx = _find_header_row(df_all)

    if header_row_idx is None:
        # No recognisable header found — attempt plain read (column check will fail gracefully).
        df = pd.read_excel(io.BytesIO(raw), converters=converters)
        return df, 2

    # Rows before the header are title banners / metadata — skip them.
    rows_to_skip = list(range(header_row_idx))

    # If the row immediately after the header is all-text (hints/description), skip it too.
    hint_skipped = False
    hint_idx = header_row_idx + 1
    if hint_idx < len(df_all):
        hint_vals = [
            v for v in df_all.iloc[hint_idx]
            if v is not None and not (isinstance(v, float) and pd.isna(v))
        ]
        if hint_vals and not any(isinstance(v, (int, float)) for v in hint_vals):
            rows_to_skip.append(hint_idx)
            hint_skipped = True

    # data_start_row (1-based Excel): header row + 1 (for header itself) + hint + 1 (1-indexed)
    data_start_row = header_row_idx + 2 + (1 if hint_skipped else 0)

    read_kwargs: dict = {
        "converters": converters,
        "na_values": ["", "N/A", "n/a", "None", "-"],
        "keep_default_na": True,
    }
    if rows_to_skip:
        read_kwargs["skiprows"] = rows_to_skip

    df = pd.read_excel(io.BytesIO(raw), **read_kwargs)
    return df, data_start_row


def _coerce_numeric_columns(df: pd.DataFrame, data_start_row: int) -> tuple[pd.DataFrame, List[str]]:
    """Coerce numeric columns with ``pd.to_numeric(errors='coerce')``.

    Returns *(modified_df, warnings)* where *warnings* is a list of
    human-readable strings for cells whose values could not be parsed as
    numbers (they are replaced with NaN / missing rather than crashing).
    """
    warnings: List[str] = []
    numeric_cols = {
        "price_selling": "Selling Price",
        "price_buying": "Buying Price",
        "stock_quantity": "Current Stock",
        "min_stock_alert": "Low Stock Limit",
        "wholesale_price": "Wholesale Price",
        "wholesale_threshold": "Wholesale Threshold",
    }
    for field, label in numeric_cols.items():
        if field not in df.columns:
            continue
        original = df[field].copy()
        df[field] = pd.to_numeric(df[field], errors="coerce")
        # Find cells that were non-null before coercion but became NaN after.
        bad_mask = original.notna() & df[field].isna()
        for idx in df.index[bad_mask]:
            excel_row = int(idx) + data_start_row
            warnings.append(
                f"Row {excel_row}: '{label}' value '{original.at[idx]}' "
                f"is not a valid number — treated as missing"
            )
    return df, warnings


def _check_intra_file_duplicates(df: pd.DataFrame, data_start_row: int) -> List[str]:
    """Return warnings for barcodes that appear more than once in the upload.

    The last occurrence of each duplicate wins on DB upsert, but all affected
    rows are surfaced in the warnings so the user knows what happened.
    """
    if "barcode" not in df.columns:
        return []
    # keep=False marks every occurrence of a duplicated value (not just 2nd+).
    dup_mask = df.duplicated(subset=["barcode"], keep=False)
    if not dup_mask.any():
        return []

    warnings: List[str] = []
    dup_df = df[dup_mask][["barcode"]].copy()
    dup_df["_excel_row"] = [int(i) + data_start_row for i in dup_df.index]
    grouped = dup_df.groupby("barcode")["_excel_row"].apply(list)
    for barcode_val, rows in grouped.items():
        rows_str = ", ".join(str(r) for r in rows)
        warnings.append(
            f"Duplicate barcode '{barcode_val}' at rows {rows_str} "
            f"— last occurrence imported"
        )
    return warnings


@router.post("/upload")
def upload_inventory(file: UploadFile = File(...)):
    """
    Bulk upload products from .xlsx or .csv.
    Required columns (any case): Item Name, Barcode, Selling Price.
    Optional: Buying Price, Current Stock, Low Stock Limit, Wholesale Price, Wholesale Threshold.
    Existing barcode → update; new barcode → create.
    Response includes per-row warnings for bad data and duplicate barcodes.
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename")
    if not (
        file.filename.lower().endswith(".xlsx")
        or file.filename.lower().endswith(".csv")
    ):
        raise HTTPException(status_code=400, detail="File must be .xlsx or .csv")

    try:
        df, data_start_row = _read_upload(file)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid file: {e}") from e

    df = _normalize_columns(df)

    # Drop fully-empty rows (blank rows in the middle of a spreadsheet)
    df = df.dropna(how="all").reset_index(drop=True)

    required = {"name", "barcode", "price_selling"}
    missing = required - set(df.columns)
    if missing:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Missing columns: {sorted(missing)}. "
                "Expected: name (or 'Item Name'), barcode, price_selling (or 'Selling Price'); "
                "optional: price_buying, stock_quantity, min_stock_alert, "
                "wholesale_price, wholesale_threshold"
            ),
        )

    # Default price_buying from price_selling when column is absent (API/test compatibility)
    if "price_buying" not in df.columns:
        df["price_buying"] = df["price_selling"]

    # Vectorised numeric coercion — bad cells become NaN with per-row warnings
    df, coerce_warnings = _coerce_numeric_columns(df, data_start_row)

    # Warn about barcodes that appear more than once in the file
    dup_warnings = _check_intra_file_duplicates(df, data_start_row)

    created = 0
    updated = 0
    errors: List[str] = list(coerce_warnings) + list(dup_warnings)

    with Session(engine) as session:
        for idx, row in df.iterrows():
            excel_row = int(idx) + data_start_row  # type: ignore[arg-type]
            barcode = str(row["barcode"]).strip()
            if not barcode or barcode.lower() in ("nan", "none", ""):
                errors.append(f"Row {excel_row}: missing barcode — skipped")
                continue
            try:
                name_val = row["name"]
                name = str(name_val).strip() if not pd.isna(name_val) else ""
                name = name or "Unknown"

                if pd.isna(row["price_selling"]):
                    errors.append(f"Row {excel_row} ({barcode}): missing Selling Price — skipped")
                    continue
                price_selling = float(row["price_selling"])
                if price_selling < 0:
                    errors.append(f"Row {excel_row} ({barcode}): Selling Price {price_selling} is negative — skipped")
                    continue

                _pb = row.get("price_buying")
                price_buying = float(_pb) if _pb is not None and not pd.isna(_pb) else price_selling
                if price_buying < 0:
                    price_buying = price_selling  # fallback rather than hard-fail

                _sq = row.get("stock_quantity")
                stock_quantity = int(_sq) if _sq is not None and not pd.isna(_sq) else 0

                _ms = row.get("min_stock_alert")
                min_stock_alert = int(_ms) if _ms is not None and not pd.isna(_ms) else 5

                _wp = row.get("wholesale_price")
                wholesale_price: Optional[float] = None
                if _wp is not None and not pd.isna(_wp):
                    try:
                        v = float(_wp)
                        wholesale_price = v if v >= 0 else None
                    except (TypeError, ValueError):
                        pass

                _wt = row.get("wholesale_threshold")
                wholesale_threshold: Optional[int] = None
                if _wt is not None and not pd.isna(_wt):
                    try:
                        v = int(_wt)
                        wholesale_threshold = v if v >= 0 else None
                    except (TypeError, ValueError):
                        pass

                existing = session.exec(
                    select(Product).where(Product.barcode == barcode)
                ).first()
                if existing:
                    existing.name = name
                    existing.price_buying = price_buying
                    existing.price_selling = price_selling
                    existing.stock_quantity = stock_quantity
                    existing.min_stock_alert = min_stock_alert
                    existing.wholesale_price = wholesale_price
                    existing.wholesale_threshold = wholesale_threshold
                    session.add(existing)
                    updated += 1
                else:
                    product = Product(
                        name=name,
                        barcode=barcode,
                        price_buying=price_buying,
                        price_selling=price_selling,
                        stock_quantity=stock_quantity,
                        min_stock_alert=min_stock_alert,
                        wholesale_price=wholesale_price,
                        wholesale_threshold=wholesale_threshold,
                    )
                    session.add(product)
                    created += 1
            except Exception as e:
                errors.append(f"Row {excel_row} ({barcode}): {e}")
                continue
        session.commit()

    return {
        "created": created,
        "updated": updated,
        "errors": errors,
    }
